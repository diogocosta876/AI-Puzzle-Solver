import os
from time import time
from collections import deque
from math import inf

import openpyxl
import pandas as pd
from matplotlib import pyplot as plt

from benchmark_levels import LEVELS

DIRECTIONS = {
    "up": (-1, 0),
    "left": (0, -1),
    "down": (1, 0),
    "right": (0, 1),
}


class TreeNode:
    def __init__(self, state, parent=None):
        self.state = state
        self.parent = parent
        self.children = []

    def add_child(self, child_node):
        self.children.append(child_node)
        child_node.parent = self

    def __hash__(self):
        return hash((tuple(tuple(inner_list) for inner_list in self.state)))


class MutualFunction:
    def __init__(self):
        self.tree_nodes = []

    def check_neighbor_piece(self, board, coordinates, cluster, direction):
        y, x = coordinates
        delta_y, delta_x = direction
        new_x = max(0, x + delta_x)
        new_y = max(0, y + delta_y)

        try:
            if board[new_y][new_x] == board[y][x] and cluster[new_y][new_x] == 0:
                cluster[new_y][new_x] = 1
                for new_direction in DIRECTIONS.values():
                    self.check_neighbor_piece(board, (new_y, new_x), cluster, new_direction)
        except:
            return None

    def get_cluster(self, board, coordinates):
        if board[coordinates[0]][coordinates[1]] == None:
            return []
        cluster = [[0 for _ in range(len(board))] for _ in range(len(board))]
        cluster[coordinates[0]][coordinates[1]] = 1
        for direction in DIRECTIONS.values():
            self.check_neighbor_piece(board, coordinates, cluster, direction)

        return cluster

    def move(self, board, coordinates, direction):

        side_size = len(board)

        cluster = self.get_cluster(board, coordinates)
        if cluster == []:
            return None

        new_board = [[None for _ in range(side_size)] for _ in range(side_size)]
        for i in range(0, side_size):
            for j in range(0, side_size):
                if cluster[i][j] == 1:
                    new_board[i][j] = None
                else:
                    new_board[i][j] = board[i][j]

        if direction == "up":
            for i in range(0, side_size):
                if cluster[0][i] == 1:
                    return None

            cluster.pop(0)
            line = [0 for _ in range(side_size)]
            cluster.append(line)

        elif direction == "down":
            for i in range(0, side_size):
                if cluster[side_size - 1][i] == 1:
                    return None
            for i in range(side_size - 1, -1, -1):
                for j in range(0, side_size):
                    cluster[i][j] = cluster[i - 1][j]
            for i in range(0, side_size):
                cluster[0][i] = 0

        elif direction == "left":
            for i in range(0, side_size):
                if cluster[i][0] == 1:
                    return None
            for i in range(0, side_size):
                for j in range(0, side_size - 1):
                    cluster[i][j] = cluster[i][j + 1]
            for i in range(0, side_size):
                cluster[i][side_size - 1] = 0

        elif direction == "right":
            for i in range(0, side_size):
                if cluster[i][side_size - 1] == 1:
                    return None
            for i in range(0, side_size):
                for j in range(side_size - 1, -1, -1):
                    cluster[i][j] = cluster[i][j - 1]
            for i in range(0, side_size):
                cluster[i][0] = 0

        color = board[coordinates[0]][coordinates[1]]

        for i in range(0, side_size):
            for j in range(0, side_size):
                if cluster[i][j] == 1:
                    if new_board[i][j] == None:
                        new_board[i][j] = color
                    else:
                        return None

        return new_board

    def get_number_clusters(self, board):
        aux_board = []
        viewed_clusters = 0

        for i in range(len(board)):
            for j in range(len(board)):
                if board[i][j] != None:
                    aux_board.append((i, j))

        while aux_board:
            i, j = aux_board[0]  # get first element
            cluster = self.get_cluster(board, (i, j))
            if cluster == []: continue
            viewed_clusters += 1
            for a, row in enumerate(cluster):
                for b, _ in enumerate(row):
                    if cluster[a][b] == 1:
                        aux_board.remove((a, b))

        return viewed_clusters

    def get_number_colors(self, board):
        elements = set()
        for line in board:
            for element in line:
                elements.add(element)

        return elements.__len__() - 1

    def child_states(self, board):
        new_states = []
        aux_board = []

        for i in range(len(board)):
            for j in range(len(board)):
                if board[i][j] != None:
                    aux_board.append((i, j))

        while aux_board:
            i, j = aux_board[0]  # get first element
            cluster = self.get_cluster(board, (i, j))
            for a, row in enumerate(cluster):
                for b, _ in enumerate(row):
                    if cluster[a][b] == 1:
                        aux_board.remove((a, b))

            for direction in DIRECTIONS.keys():
                next_board = self.move(board, (i, j), direction)
                if next_board:
                    new_states.append(next_board)

        return new_states

    def win(self, board):
        return self.get_number_clusters(board) == self.get_number_colors(board)

    def get_steps(self, solution: TreeNode):
        steps_counter = -1
        steps = []
        while solution:
            steps.append(solution.state)
            steps_counter += 1
            solution = solution.parent
        steps.reverse()

        return steps_counter, steps

    def depth(self, node) -> int:
        depth = 0
        while node.parent is not None:
            node = node.parent
            depth += 1
        return depth

    def timed_out(self, before) -> bool:
        return time() - before > 60

    def search_heuristic_1(self, board):
        return self.get_number_clusters(board) - self.get_number_colors(board)

    def search_heuristic_2(self, board):
        clusters = []
        for (i, j) in [(i, j) for i in range(len(board)) for j in range(len(board))]:
            cluster = self.get_cluster(board, [i, j])
            if cluster == []:
                continue
            if (cluster, board[i][j]) not in clusters:
                clusters.append((cluster, board[i][j]))

        total = 0
        for i1 in range(0, len(clusters) - 1):
            for i2 in range(i1 + 1, len(clusters)):
                (cluster1, color1), (cluster2, color2) = clusters[i1], clusters[i2]
                if color1 == color2:
                    total += self.get_distance(cluster1, cluster2)

        return total

    def search_heuristic_3(self, board):
        return self.a_star_search_heuristic1(board) * len(board)

class ASTAR(MutualFunction):
    def __init__(self):
        super().__init__()

    def run(self, board: list, score: list, heuristic_index: int):
        self.before = time()

        if heuristic_index == 1:
            nodes, visits, mem = self.a_star_search(board, self.win, self.child_states, self.a_star_search_heuristic1)
        elif heuristic_index == 2:
            nodes, visits, mem = self.a_star_search(board, self.win, self.child_states, self.a_star_search_heuristic2)
        else:
            nodes, visits, mem = self.a_star_search(board, self.win, self.child_states, self.a_star_search_heuristic3)

        if nodes is None:
            score[0] = "N/A"
            score[1] = "N/A"
        else:
            self.tree_nodes = nodes
            score[0] += round(time() - self.before, 2)  # Time
            score[1] = self.get_steps(nodes)[0]  # Moves and Board states for each move
            score[5] = True  # Solved
            score[6] = visits
            score[7] = mem

        return score

    def shortest_path(self, initial_coordinate, cluster):
        visited = set()
        queue = deque([(initial_coordinate, [])])

        while queue:
            (row, col), path = queue.popleft()
            if (row, col) in visited:
                continue
            visited.add((row, col))

            if cluster[row][col] == 1:
                return len(path)

            for vx, vy in DIRECTIONS.values():
                next_x, next_y = row + vx, col + vy
                if 0 <= next_x < len(cluster) and 0 <= next_y < len(cluster) and (next_x, next_y) not in visited:
                    queue.append(((next_x, next_y), path + [(row, col)]))

        return None

    def get_distance(self, cluster1, cluster2):
        total = inf
        for (i, j) in [(i, j) for i in range(len(cluster1)) for j in range(len(cluster1))]:
            if cluster1[i][j] == 1:
                total = min(self.shortest_path((i, j), cluster2), total)
        return total

    def a_star_search_heuristic1(self, board):
        return self.get_number_clusters(board) - self.get_number_colors(board)

    def a_star_search_heuristic2(self, board):
        clusters = []
        for (i, j) in [(i, j) for i in range(len(board)) for j in range(len(board))]:
            cluster = self.get_cluster(board, [i, j])
            if cluster == []:
                continue
            if (cluster, board[i][j]) not in clusters:
                clusters.append((cluster, board[i][j]))

        total = 0
        for i1 in range(0, len(clusters) - 1):
            for i2 in range(i1 + 1, len(clusters)):
                (cluster1, color1), (cluster2, color2) = clusters[i1], clusters[i2]
                if color1 == color2:
                    total += self.get_distance(cluster1, cluster2)

        return total

    def a_star_search_heuristic3(self, board):
        return self.a_star_search_heuristic1(board) * len(board)

    def a_star_search(self, initial_state, goal_state_func, operators_func, heuristic_func):

        root = TreeNode(initial_state)  # create the root node in the search tree
        stack = [(root, heuristic_func(initial_state))]  # initialize the queue to store the nodes
        filtered_states = [initial_state]
        visited_nodes = 1  # Count the root node as visited
        max_nodes_in_memory = 1  # Keep track of the maximum number of nodes in memory

        while len(stack):

            if self.timed_out(self.before): return None, visited_nodes, max_nodes_in_memory

            node, _ = stack.pop()  # get first element in the queue
            # print("nó com valor", v)
            if goal_state_func(node.state):  # check goal state
                return node, visited_nodes, max_nodes_in_memory

            children = operators_func(node.state)
            evaluated_children = [(child, heuristic_func(child) + self.depth(node) + 1) for child in children]

            for (child, value) in evaluated_children:  # go through next states
                if child in filtered_states:
                    continue

                filtered_states.append(child)
                visited_nodes += 1  # Increment the visited nodes count

                # create tree node with the new state
                child_tree = TreeNode(child, node)

                node.add_child(child_tree)

                # enqueue the child node
                stack.append((child_tree, value))

                # Update max nodes in memory count
                max_nodes_in_memory = max(max_nodes_in_memory, len(stack))

            stack = sorted(stack, key=lambda node: node[1], reverse=True)

        return None, visited_nodes, max_nodes_in_memory


class BFS(MutualFunction):
    def __init__(self):
        super().__init__()
        self.visited_nodes = 0
        self.max_memory_usage = 0

    def run(self, board: list, score: list):
        before = time()
        (nodes, visited_nodes, max_memory_usage) = self.breadth_first_search(initial_state=board,
                                                                             goal_state_func=self.win,
                                                                             operators_func=self.child_states)
        self.visited_nodes = visited_nodes
        self.max_memory_usage = max_memory_usage
        score[0] += round(time() - before, 2)  # Time
        score[1], score[3] = self.get_steps(nodes)  # Moves and Board states for each move

    def breadth_first_search(self, initial_state, goal_state_func, operators_func):
        root = TreeNode(initial_state)  # create the root node in the search tree
        queue = deque([root])  # initialize the queue to store the nodes
        visited = [initial_state]
        visited_nodes = 1  # initialize the visited nodes counter to 1 (for the root node)
        max_memory_usage = 1  # initialize the maximum memory usage to 1 (for the root node)

        while queue:  # while the queue is not empty
            node = queue.popleft()  # get first element in the queue
            if goal_state_func(node.state):  # check goal state
                return node, visited_nodes, max_memory_usage

            for state in operators_func(node.state):
                if state not in visited:  # go through next states
                    child = TreeNode(state=state, parent=node)
                    node.add_child(child)
                    queue.append(child)
                    visited.append(state)
                    visited_nodes += 1  # increment visited nodes counter
                    max_memory_usage = max(max_memory_usage, len(queue) + visited_nodes)  # update maximum memory usage

        return None, visited_nodes, max_memory_usage  # no solution found


class DFS(MutualFunction):
    def __init__(self):
        super().__init__()
        self.visited_nodes = 0
        self.max_memory_usage = 0

    def run(self, board: list, score: list):
        before = time()
        (nodes, visited_nodes, max_memory_usage) = self.depth_first_search(board, self.win, self.child_states)
        self.visited_nodes = visited_nodes
        self.max_memory_usage = max_memory_usage
        score[0] += round(time() - before, 2)  # Time
        score[1], score[3] = self.get_steps(nodes)  # Moves and Board states for each move

    def depth_first_search(self, initial_state, goal_state_func, operators_func):
        root = TreeNode(initial_state)  # create the root node in the search tree
        stack = [root]  # initialize the stack to store the nodes
        filtered_states = []
        max_memory = 0
        visits = 0

        while stack:  # while the stack is not empty
            node = stack.pop()  # get last element in the stack
            visits += 1
            if goal_state_func(node.state):  # check goal state
                return node, visits, max_memory

            for state in operators_func(node.state):  # go through next states
                if state in filtered_states:
                    continue
                else:
                    filtered_states.append(state)

                # create tree node with the new state
                child_tree = TreeNode(state, node)

                # link child node to its parent in the tree
                node.add_child(child_tree)

                # push the child node to the stack
                stack.append(child_tree)
                memory_usage = len(stack)
                if memory_usage > max_memory:
                    max_memory = memory_usage

        return None, visits, max_memory  # no solution found


class GREEDY(MutualFunction):
    def __init__(self):
        super().__init__()
        self.visited_nodes = 0
        self.max_memory_usage = 0

    def run(self, board: list, score: list) -> None:  # The result is stored in the score list
        before = time()
        (nodes, visited_nodes, max_memory_usage) = self.greedy_search(board, self.win, self.child_states,
                                                                      self.greedy_heuristic)
        self.visited_nodes = visited_nodes
        self.max_memory_usage = max_memory_usage
        score[0] += round(time() - before, 2)  # Time
        score[1], score[3] = self.get_steps(nodes)  # Moves and Board states for each move

    def greedy_heuristic(self, board):
        return self.get_number_clusters(board) - self.get_number_colors(board)

    def greedy_search(self, initial_state, goal_state_func, operators_func, heuristic_func):
        num_visits = 0
        max_memory = 0

        root = TreeNode(initial_state)  # create the root node in the search tree
        stack = [(root, heuristic_func(initial_state))]  # initialize the queue to store the nodes
        filtered_states = [initial_state]

        while len(stack):

            num_visits += 1
            max_memory = max(max_memory, len(filtered_states))

            node, v = stack.pop()  # get first element in the queue
            if goal_state_func(node.state):  # check goal state
                return node, num_visits, max_memory

            children = operators_func(node.state)
            evaluated_children = [(child, heuristic_func(child)) for child in children]

            for (child, value) in evaluated_children:  # go through next states
                if child in filtered_states:
                    continue

                filtered_states.append(child)

                # create tree node with the new state
                child_tree = TreeNode(child, node)

                node.add_child(child_tree)

                # enqueue the child node
                stack.append((child_tree, value))

            stack = sorted(stack, key=lambda node: node[1], reverse=True)

        return None, num_visits, max_memory


def analysis():

    # CONFIGURATIONS

    RUN_DFS = True
    RUN_BFS = True
    RUN_GREEDY = True
    RUN_ASTAR = True

    TIME_DFS = 0
    TIME_BFS = 0
    TIME_GREEDY = 0
    TIME_ASTAR = 0

    MOVES_DFS = 0
    MOVES_BFS = 0
    MOVES_GREEDY = 0
    MOVES_ASTAR = 0

    file_path = "results.xlsx"

    if os.path.isfile(file_path):
        # If the file exists, open it and append to the existing sheets
        workbook = openpyxl.load_workbook(file_path)
    else:
        workbook = openpyxl.Workbook()


    sheet_name = 'Data'
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(sheet_name)

    # Add column headers
    worksheet.cell(row=1, column=1, value="LEVEL")
    worksheet.cell(row=1, column=2, value="DFS_TIME")
    worksheet.cell(row=1, column=3, value="BFS_TIME")
    worksheet.cell(row=1, column=4, value="GREEDY_TIME")
    worksheet.cell(row=1, column=5, value="ASTAR_H1_TIME")
    worksheet.cell(row=1, column=6, value="ASTAR_H2_TIME")
    worksheet.cell(row=1, column=7, value="ASTAR_H3_TIME")
    worksheet.cell(row=1, column=8, value="DFS_MOVES")
    worksheet.cell(row=1, column=9, value="BFS_MOVES")
    worksheet.cell(row=1, column=10, value="GREEDY_MOVES")
    worksheet.cell(row=1, column=11, value="ASTAR_H1_MOVES")
    worksheet.cell(row=1, column=12, value="ASTAR_H2_MOVES")
    worksheet.cell(row=1, column=13, value="ASTAR_H3_MOVES")
    worksheet.cell(row=1, column=14, value="DFS_VISITS")
    worksheet.cell(row=1, column=15, value="BFS_VISITS")
    worksheet.cell(row=1, column=16, value="GREEDY_VISITS")
    worksheet.cell(row=1, column=17, value="ASTAR_H2_VISITS")
    worksheet.cell(row=1, column=18, value="DFS_MAX_MEM")
    worksheet.cell(row=1, column=19, value="BFS_MAX_MEM")
    worksheet.cell(row=1, column=21, value="GREEDY_MAX_MEM")
    worksheet.cell(row=1, column=22, value="ASTAR_H2_MAX_MEM")

    row = 1
    for LEVEL in LEVELS:
        print("level", LEVEL)
        row += 1
        worksheet.cell(row=row, column=1, value=LEVEL)

        if RUN_DFS:
            print("Depth-first search algorithm")
            dfs = DFS()
            score = [0.0, None, None, None]
            dfs.run(LEVELS[LEVEL], score)
            TIME_DFS = score[0]
            MOVES_DFS = score[1]
            NODE_VISITS = dfs.visited_nodes
            MAX_MEM = dfs.max_memory_usage
            print("Level: ", LEVEL, " Time: ", TIME_DFS, " Moves: ", MOVES_DFS)
            print("node visits:", NODE_VISITS, "  max mem usage:", MAX_MEM, '\n')
            worksheet.cell(row=row, column=2, value=TIME_DFS)
            worksheet.cell(row=row, column=8, value=MOVES_DFS)
            worksheet.cell(row=row, column=14, value=NODE_VISITS)
            worksheet.cell(row=row, column=18, value=MAX_MEM)

        if RUN_BFS:
            print("Breath-first search algorithm")
            bfs = BFS()
            score = [0.0, None, None, None]
            bfs.run(LEVELS[LEVEL], score)
            TIME_BFS = score[0]
            MOVES_BFS = score[1]
            NODE_VISITS = bfs.visited_nodes
            MAX_MEM = bfs.max_memory_usage
            print("Level: ", LEVEL, " Time: ", TIME_BFS, " Moves: ", MOVES_BFS)
            print("node visits:", NODE_VISITS, "  max mem usage:", MAX_MEM, '\n')
            worksheet.cell(row=row, column=3, value=TIME_BFS)
            worksheet.cell(row=row, column=9, value=MOVES_BFS)
            worksheet.cell(row=row, column=15, value=NODE_VISITS)
            worksheet.cell(row=row, column=19, value=MAX_MEM)

        if RUN_GREEDY:
            print("Greedy search algorithm")
            greedy = GREEDY();
            score = [0.0, None, None, None, None]
            greedy.run(LEVELS[LEVEL], score)
            TIME_GREEDY = score[0]
            MOVES_GREEDY = score[1]
            NODE_VISITS = greedy.visited_nodes
            MAX_MEM = greedy.max_memory_usage
            print("Level: ", LEVEL, " Time: ", TIME_GREEDY, "Moves: ", MOVES_GREEDY)
            print("node visits:", NODE_VISITS, "  memory spent:", MAX_MEM, '\n')
            worksheet.cell(row=row, column=4, value=TIME_GREEDY)
            worksheet.cell(row=row, column=10, value=MOVES_GREEDY)
            worksheet.cell(row=row, column=16, value=NODE_VISITS)  # visits
            worksheet.cell(row=row, column=21, value=MAX_MEM)  # max memory utilized

        if RUN_ASTAR:
            print("A Star search algorithm")
            astar = ASTAR()
            for index in range(0,3):
                score = [0.0, None, None, None, None, None, None, None, None]
                level = LEVELS[LEVEL]
                astar.run(level, score, index)
                TIME_ASTAR = score[0]
                MOVES_ASTAR = score[1]
                NODE_VISITS = score[6]
                MAX_MEM = score[7]
                # MOVES_ASTAR
                print("Heuristic: ", index, "Level: ", LEVEL, " Time: ", TIME_ASTAR, "Moves: ", MOVES_ASTAR)
                print("node visits:", NODE_VISITS, "  memory spent:", MAX_MEM, '\n')
                worksheet.cell(row=row, column=5 + index, value=TIME_ASTAR)  # TIME_ASTAR
                worksheet.cell(row=row, column=11 + index, value=MOVES_ASTAR)  # MOVES_ASTAR
                if index == 1:
                    worksheet.cell(row=row, column=17, value=NODE_VISITS)  # visits solution[1]
                    worksheet.cell(row=row, column=22, value=MAX_MEM)  # max memory utilized solution[2]

                # show_steps(solution)
                # show_graph(get_root(solution))

        workbook.save("results.xlsx")


def create_graphs():
    # Load the results file into a pandas DataFrame
    results_df = pd.read_excel('results.xlsx', sheet_name='Data')

    # Calculate the mean values of the columns you are interested in
    mean_moves = results_df[['DFS_MOVES', 'BFS_MOVES', 'GREEDY_MOVES', 'ASTAR_H2_MOVES']].mean()

    plt.figure()
    # Plot the mean values as a bar graph
    ax1 = mean_moves.plot.bar(figsize=(10, 5))
    ax1.set_xlabel('Algorithms')
    ax1.set_ylabel('Mean Move Count')
    ax1.set_title('Mean Move Count Across All Levels')

    # Update the x-axis labels
    new_labels = ['DFS', 'BFS', 'GREEDY', 'A*']
    ax1.set_xticklabels(new_labels, rotation=0)

    # Calculate the mean values of the columns you are interested in
    mean_time = results_df[['DFS_TIME', 'BFS_TIME', 'GREEDY_TIME', 'ASTAR_H2_TIME']].mean()

    plt.figure()
    # Plot the mean values as a bar graph
    ax2 = mean_time.plot.bar(figsize=(10, 5))
    ax2.set_xlabel('Algorithms')
    ax2.set_ylabel('Execution Time (s)')
    ax2.set_title('Mean Time of Execution Across All Levels')
    ax2.set_xticklabels(new_labels, rotation=0)

    # Calculate the mean values of the columns you are interested in
    mean_visits = results_df[['DFS_VISITS', 'BFS_VISITS', 'GREEDY_VISITS', 'ASTAR_H2_VISITS']].mean()

    plt.figure()
    # Plot the mean values as a bar graph
    ax3 = mean_visits.plot.bar(figsize=(10, 5))
    ax3.set_xlabel('Algorithms')
    ax3.set_ylabel('Node Visits')
    ax3.set_title('Mean Node Visits Across All Levels')
    ax3.set_xticklabels(new_labels, rotation=0)

    mean_mem = results_df[['DFS_MAX_MEM', 'BFS_MAX_MEM', 'GREEDY_MAX_MEM', 'ASTAR_H2_MAX_MEM']].mean()

    plt.figure()
    # Plot the mean values as a bar graph
    ax4 = mean_mem.plot.bar(figsize=(10, 5))
    ax4.set_xlabel('Algorithms')
    ax4.set_ylabel('Nodes Stored in Memory')
    ax4.set_title('Mean Max Memory Usage Across All Levels')
    ax4.set_xticklabels(new_labels, rotation=0)

    mean_times = results_df[['ASTAR_H1_TIME', 'ASTAR_H2_TIME', 'ASTAR_H3_TIME']].mean()

    new_labels2 = ['A* Heuristic 1', 'A* Heuristic 2', 'A* Heuristic 3']
    plt.figure()
    # Plot the mean values as a bar graph
    ax5 = mean_times.plot.bar(figsize=(10, 5))
    ax5.set_xlabel('Algorithms')
    ax5.set_ylabel('Execution Time (s)')
    ax5.set_title('Mean Time of Execution Across All Levels')
    ax5.set_xticklabels(new_labels2, rotation=0)

    mean_moves = results_df[['ASTAR_H1_MOVES', 'ASTAR_H2_MOVES', 'ASTAR_H3_MOVES']].mean()

    plt.figure()
    # Plot the mean values as a bar graph
    ax6 = mean_moves.plot.bar(figsize=(10, 5))
    ax6.set_xlabel('Algorithms')
    ax6.set_ylabel('Mean Move Count')
    ax6.set_title('Mean Move Count Across All Levels')
    ax6.set_xticklabels(new_labels2, rotation=0)

    plt.show()

if __name__ == "__main__":
    #analysis()
    create_graphs()
